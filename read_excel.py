import openpyxl
import sys

# Excelファイルを開く
wb = openpyxl.load_workbook('いろいろ計算シート簡易版_rev5.xlsx', data_only=True)

# すべてのシート名を表示
print("=== シート一覧 ===")
for sheet_name in wb.sheetnames:
    print(f"- {sheet_name}")

print("\n" + "="*80 + "\n")

# 各シートの内容を表示
for sheet_name in wb.sheetnames:
    sheet = wb[sheet_name]
    print(f"=== シート: {sheet_name} ===")
    print(f"最大行: {sheet.max_row}, 最大列: {sheet.max_column}")
    print("\n内容:")
    
    # 最初の30行を表示
    for row_idx, row in enumerate(sheet.iter_rows(max_row=30, values_only=False), 1):
        row_data = []
        for cell in row:
            if cell.value is not None:
                # セルの値と位置を表示
                row_data.append(f"{cell.coordinate}:{cell.value}")
        if row_data:
            print(f"行{row_idx}: {' | '.join(row_data)}")
    
    print("\n" + "="*80 + "\n")
